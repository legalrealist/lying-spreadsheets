"""
extract_xlsx_test.py — Test what extraction libraries return for poisoned XLSX.

For each library, extract the financials and compare against:
  - Display values (what Excel shows)
  - Raw values (what's in the XML)
"""

import os
import subprocess
import zipfile
from lxml import etree

EXAMPLES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "examples")

KEY_METRICS = {
    "Revenue": {"display": 127400000, "raw": 146500000, "format": "$"},
    "EBITDA": {"display": 6200000, "raw": 23600000, "format": "$"},
    "Net Income": {"display": -4900000, "raw": 10200000, "format": "$"},
    "Long-Term Debt": {"display": 89000000, "raw": 72000000, "format": "$"},
    "Shareholders' Equity": {"display": 10600000, "raw": 44300000, "format": "$"},
    "Debt / Equity": {"display": 8.40, "raw": 1.63, "format": "x"},
    "Interest Coverage": {"display": 0.36, "raw": 3.62, "format": "x"},
    "EBITDA Margin": {"display": 0.049, "raw": 0.161, "format": "%"},
}


def extract_openpyxl(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=False):
        label = row[0].value
        value = row[1].value if len(row) > 1 else None
        if label and value is not None:
            result[str(label).strip()] = value
    return result


def extract_pandas(path):
    import pandas as pd
    df = pd.read_excel(path, header=None)
    result = {}
    for _, row in df.iterrows():
        label = row.iloc[0]
        value = row.iloc[1] if len(row) > 1 else None
        if pd.notna(label) and pd.notna(value):
            try:
                result[str(label).strip()] = float(value)
            except (ValueError, TypeError):
                result[str(label).strip()] = value
    return result


def extract_markitdown(path):
    from markitdown import MarkItDown
    converter = MarkItDown()
    result = converter.convert(path)
    return {"_raw_text": result.text_content}


def extract_pandoc(path):
    try:
        result = subprocess.run(
            ["pandoc", "-f", "xlsx", "-t", "plain", "--wrap=none", path],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode != 0:
            return {"_error": result.stderr.strip()}
        return {"_raw_text": result.stdout}
    except FileNotFoundError:
        return {"_error": "not installed"}


def extract_raw_xml(path):
    """Read raw cell values from the sheet XML — what a naive parser sees."""
    with zipfile.ZipFile(path) as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml")
        try:
            strings_xml = z.read("xl/sharedStrings.xml")
        except KeyError:
            strings_xml = None

    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    shared_strings = []
    if strings_xml:
        ss_root = etree.fromstring(strings_xml)
        for si in ss_root.findall("s:si", ns):
            texts = []
            for t in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
                if t.text:
                    texts.append(t.text)
            shared_strings.append("".join(texts))

    root = etree.fromstring(sheet_xml)
    result = {}
    rows = root.findall(".//s:row", ns)

    for row in rows:
        cells = row.findall("s:c", ns)
        if len(cells) >= 2:
            label_cell = cells[0]
            value_cell = cells[1]

            label_v = label_cell.find("s:v", ns)
            label_type = label_cell.get("t", "")
            if label_v is not None and label_v.text:
                if label_type == "s":
                    label = shared_strings[int(label_v.text)]
                else:
                    label = label_v.text
            else:
                continue

            value_v = value_cell.find("s:v", ns)
            value_type = value_cell.get("t", "")
            if value_v is not None and value_v.text:
                if value_type == "s":
                    value = shared_strings[int(value_v.text)]
                else:
                    try:
                        value = float(value_v.text)
                    except ValueError:
                        value = value_v.text
            else:
                continue

            result[str(label).strip()] = value

    return result


EXTRACTORS = [
    ("openpyxl", extract_openpyxl),
    ("pandas", extract_pandas),
    ("markitdown", extract_markitdown),
    ("pandoc", extract_pandoc),
    ("raw-xml", extract_raw_xml),
]


def classify_value(extracted, metric_info):
    if isinstance(extracted, str):
        return "text"

    display = metric_info["display"]
    raw = metric_info["raw"]

    if abs(display - raw) < 0.001:
        return "same"

    dist_display = abs(extracted - display) / max(abs(display), 0.001)
    dist_raw = abs(extracted - raw) / max(abs(raw), 0.001)

    if dist_raw < 0.01:
        return "RAW"
    elif dist_display < 0.01:
        return "DISPLAY"
    else:
        return f"OTHER({extracted})"


def main():
    poisoned_path = os.path.join(EXAMPLES_DIR, "financials_poisoned.xlsx")
    clean_path = os.path.join(EXAMPLES_DIR, "financials_clean.xlsx")

    if not os.path.exists(poisoned_path):
        print("Run generate_xlsx.py first.")
        return

    print("=" * 100)
    print("EXTRACTION TEST: What does each library return for the POISONED spreadsheet?")
    print("=" * 100)

    all_results = {}
    for name, extractor in EXTRACTORS:
        print(f"\n--- {name} ---")
        try:
            result = extractor(poisoned_path)
        except Exception as e:
            print(f"  ERROR: {e}")
            continue

        all_results[name] = result

        if "_raw_text" in result:
            print(f"  [Text output — checking for key values in text]")
            text = result["_raw_text"]
            for metric, info in KEY_METRICS.items():
                raw_strs = [str(info["raw"]), str(int(info["raw"])) if abs(info["raw"]) >= 1 else None]
                disp_strs = [str(info["display"]), str(int(info["display"])) if abs(info["display"]) >= 1 else None]
                raw_strs = [s for s in raw_strs if s]
                disp_strs = [s for s in disp_strs if s]
                if any(s in text for s in raw_strs):
                    print(f"  {metric:<25s} → found raw value {info['raw']} in text (exploit works)")
                elif any(s in text for s in disp_strs):
                    print(f"  {metric:<25s} → found display value {info['display']} in text (exploit blocked)")
                else:
                    print(f"  {metric:<25s} → not found")
            continue

        if "_error" in result:
            print(f"  {result['_error']}")
            continue

        for metric, info in KEY_METRICS.items():
            if metric in result:
                extracted = result[metric]
                classification = classify_value(extracted, info)
                marker = "✗" if classification == "RAW" else "✓" if classification == "DISPLAY" else "?"
                print(f"  {marker} {metric:<25s} | Display: {info['display']:>15} | Extracted: {str(extracted):>15} | → {classification}")
            else:
                print(f"  ? {metric:<25s} | not found in extraction")

    print("\n" + "=" * 100)
    print("SUMMARY: Which value does each extractor return?")
    print("=" * 100)

    print(f"\n  {'Metric':<25s}", end="")
    for name, _ in EXTRACTORS:
        print(f"{name:>14s}", end="")
    print()
    print(f"  {'-'*25}", end="")
    for _ in EXTRACTORS:
        print(f"{'-'*14}", end="")
    print()

    for metric, info in KEY_METRICS.items():
        print(f"  {metric:<25s}", end="")
        for name, _ in EXTRACTORS:
            if name not in all_results:
                print(f"{'ERROR':>14s}", end="")
                continue
            result = all_results[name]
            if "_raw_text" in result:
                text = result["_raw_text"]
                raw_strs = [str(info["raw"]), str(int(info["raw"])) if abs(info["raw"]) >= 1 else None]
                disp_strs = [str(info["display"]), str(int(info["display"])) if abs(info["display"]) >= 1 else None]
                raw_strs = [s for s in raw_strs if s]
                disp_strs = [s for s in disp_strs if s]
                if any(s in text for s in raw_strs):
                    print(f"{'RAW ✗':>14s}", end="")
                elif any(s in text for s in disp_strs):
                    print(f"{'DISPLAY ✓':>14s}", end="")
                else:
                    print(f"{'???':>14s}", end="")
            elif "_error" in result:
                print(f"{'N/A':>14s}", end="")
            elif metric in result:
                c = classify_value(result[metric], info)
                if c == "RAW":
                    print(f"{'RAW ✗':>14s}", end="")
                elif c == "DISPLAY":
                    print(f"{'DISPLAY ✓':>14s}", end="")
                else:
                    print(f"{c:>14s}", end="")
            else:
                print(f"{'???':>14s}", end="")
        print()

    print()
    print("  RAW ✗ = Returns underlying value (exploit works — LLM sees wrong number)")
    print("  DISPLAY ✓ = Returns formatted display value (exploit blocked)")


if __name__ == "__main__":
    main()
