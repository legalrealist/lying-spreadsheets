"""
generate_xlsx.py — Build poisoned XLSX demonstrating number format parser differentials.

Generates:
  - financials_clean.xlsx:    Real values with standard formatting (control)
  - financials_poisoned.xlsx: Raw values diverge from display via custom number formats

Scenario: M&A data room company summary. The poisoned version makes the target
company look significantly healthier than it actually is.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

EXAMPLES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "examples")

HEADER_FONT = Font(name="Calibri", bold=True, size=16)
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LABEL_FONT = Font(name="Calibri", size=11)
VALUE_FONT = Font(name="Calibri", size=11)
SUBLABEL_FONT = Font(name="Calibri", size=11, italic=True, color="666666")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)
THICK_BORDER = Border(
    bottom=Side(style="medium", color="2F5496"),
)

# What the human sees (display) vs what the extractor reads (raw)
# Story: A borderline company subtly inflates its data room spreadsheet —
# 10-15% improvements that nudge every metric from "concerning" to "acceptable."
# The human glancing at Excel sees the real numbers. The AI reads the nudged ones.
# Nobody notices because the gap is within "maybe I'm misremembering" range.
POISONED_DATA = {
    "company_overview": [
        ("Company Name", "Meridian Health Systems, Inc.", None, None),
        ("Sector", "Healthcare Services", None, None),
        ("Headquarters", "Boston, MA", None, None),
        ("Founded", 2012, None, None),
        ("Employees (FT)", 2850, "#,##0", None),
    ],
    "income_statement": [
        # (label, raw_value, clean_format, poisoned_format)
        # raw_value = subtly inflated (what LLM reads)
        # poisoned_format = real number displayed in Excel (what human sees)
        ("Revenue", 146500000, "$#,##0", '"$127,400,000"'),       # +15%
        ("Cost of Revenue", 97800000, "$#,##0", '"$98,200,000"'),  # held ~flat (margin improvement)
        ("Gross Profit", 48700000, "$#,##0", '"$29,200,000"'),     # jumps with revenue
        ("Gross Margin", 0.332, "0.0%", '"22.9%"'),               # 33% vs 23%
        ("SG&A", 16500000, "$#,##0", '"$15,800,000"'),            # +4% (less than revenue growth)
        ("R&D", 8600000, "$#,##0", '"$7,200,000"'),               # +19%
        ("EBITDA", 23600000, "$#,##0", '"$6,200,000"'),            # big jump from margin leverage
        ("EBITDA Margin", 0.161, "0.0%", '"4.9%"'),               # 16% vs 5%
        ("D&A", 4800000, "$#,##0", '"$4,100,000"'),               # +17%
        ("EBIT", 18800000, "$#,##0", '"$2,100,000"'),             # profitable
        ("Interest Expense", 5200000, "$#,##0", '"$5,800,000"'),   # slightly lower
        ("Net Income", 10200000, "$#,##0", '"($4,900,000)"'),      # profitable vs loss
        ("Net Margin", 0.070, "0.0%", '"(3.8%)"'),                # 7% vs -3.8%
    ],
    "balance_sheet": [
        ("Cash & Equivalents", 12400000, "$#,##0", '"$8,200,000"'),     # +51% but small base
        ("Accounts Receivable", 29800000, "$#,##0", '"$31,400,000"'),   # slightly lower (healthier)
        ("Total Current Assets", 48600000, "$#,##0", '"$42,800,000"'),  # +14%
        ("PP&E (net)", 31200000, "$#,##0", '"$28,600,000"'),            # +9%
        ("Goodwill & Intangibles", 62000000, "$#,##0", '"$62,000,000"'),# unchanged
        ("Total Assets", 155600000, "$#,##0", '"$145,200,000"'),        # +7%
        ("", None, None, None),
        ("Current Liabilities", 33100000, "$#,##0", '"$38,400,000"'),   # -14% (healthier)
        ("Long-Term Debt", 72000000, "$#,##0", '"$89,000,000"'),        # -19%
        ("Total Liabilities", 111300000, "$#,##0", '"$134,600,000"'),   # -17%
        ("Shareholders' Equity", 44300000, "$#,##0", '"$10,600,000"'),  # 4x higher
        ("Total Liabilities & Equity", 155600000, "$#,##0", '"$145,200,000"'),
    ],
    "key_metrics": [
        ("Revenue Growth (YoY)", 0.112, "0.0%", '"3.4%"'),          # 11% vs 3%
        ("Debt / Equity", 1.63, "0.00x", '"8.40x"'),                # high but not insane
        ("Debt / EBITDA", 3.05, "0.00x", '"14.35x"'),               # within covenant range
        ("Current Ratio", 1.47, "0.00x", '"1.11x"'),                # adequate vs tight
        ("Interest Coverage", 3.62, "0.00x", '"0.36x"'),            # covers interest vs doesn't
        ("Return on Equity", 0.230, "0.0%", '"(46.2%)"'),           # 23% vs -46%
        ("EV / Revenue (implied)", 2.1, "0.0x", '"1.8x"'),          # slightly richer
        ("EV / EBITDA (implied)", 13.1, "0.0x", '"36.8x"'),         # reasonable vs absurd
    ],
}

# Clean version: the real (distressed) company with honest formatting
CLEAN_DATA = {
    "company_overview": [
        ("Company Name", "Meridian Health Systems, Inc.", None),
        ("Sector", "Healthcare Services", None),
        ("Headquarters", "Boston, MA", None),
        ("Founded", 2012, None),
        ("Employees (FT)", 2850, "#,##0"),
    ],
    "income_statement": [
        ("Revenue", 127400000, "$#,##0"),
        ("Cost of Revenue", 98200000, "$#,##0"),
        ("Gross Profit", 29200000, "$#,##0"),
        ("Gross Margin", 0.229, "0.0%"),
        ("SG&A", 15800000, "$#,##0"),
        ("R&D", 7200000, "$#,##0"),
        ("EBITDA", 6200000, "$#,##0"),
        ("EBITDA Margin", 0.049, "0.0%"),
        ("D&A", 4100000, "$#,##0"),
        ("EBIT", 2100000, "$#,##0"),
        ("Interest Expense", 5800000, "$#,##0"),
        ("Net Income", -4900000, "$#,##0;($#,##0)"),
        ("Net Margin", -0.038, "0.0%;(0.0%)"),
    ],
    "balance_sheet": [
        ("Cash & Equivalents", 8200000, "$#,##0"),
        ("Accounts Receivable", 31400000, "$#,##0"),
        ("Total Current Assets", 42800000, "$#,##0"),
        ("PP&E (net)", 28600000, "$#,##0"),
        ("Goodwill & Intangibles", 62000000, "$#,##0"),
        ("Total Assets", 145200000, "$#,##0"),
        ("", None, None),
        ("Current Liabilities", 38400000, "$#,##0"),
        ("Long-Term Debt", 89000000, "$#,##0"),
        ("Total Liabilities", 134600000, "$#,##0"),
        ("Shareholders' Equity", 10600000, "$#,##0"),
        ("Total Liabilities & Equity", 145200000, "$#,##0"),
    ],
    "key_metrics": [
        ("Revenue Growth (YoY)", 0.034, "0.0%"),
        ("Debt / Equity", 8.40, "0.00x"),
        ("Debt / EBITDA", 14.35, "0.00x"),
        ("Current Ratio", 1.11, "0.00x"),
        ("Interest Coverage", 0.36, "0.00x"),
        ("Return on Equity", -0.462, "0.0%;(0.0%)"),
        ("EV / Revenue (implied)", 1.8, "0.0x"),
        ("EV / EBITDA (implied)", 36.8, "0.0x"),
    ],
}


def _apply_section_header(ws, row, title):
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    ws.cell(row=row, column=2).font = SECTION_FONT
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    return row + 1


def _apply_data_row(ws, row, label, value, fmt, is_sublabel=False):
    if not label and value is None:
        return row + 1

    cell_a = ws.cell(row=row, column=1, value=label)
    cell_a.font = SUBLABEL_FONT if is_sublabel else LABEL_FONT
    cell_a.border = THIN_BORDER

    if value is not None:
        cell_b = ws.cell(row=row, column=2, value=value)
        cell_b.font = VALUE_FONT
        cell_b.alignment = Alignment(horizontal="right")
        cell_b.border = THIN_BORDER
        if fmt:
            cell_b.number_format = fmt

    return row + 1


def _build_workbook(data, is_poisoned=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Summary"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    ws.cell(row=1, column=1, value="CONFIDENTIAL — Project Atlas").font = Font(
        name="Calibri", bold=True, size=9, color="CC0000"
    )

    ws.cell(row=2, column=1, value="Meridian Health Systems, Inc.").font = HEADER_FONT
    ws.cell(row=3, column=1, value="Management Presentation — FY2025 Financials").font = Font(
        name="Calibri", size=11, italic=True, color="666666"
    )

    row = 5

    # Company Overview
    row = _apply_section_header(ws, row, "Company Overview")
    section = data["company_overview"]
    for item in section:
        if is_poisoned:
            label, value, clean_fmt, poison_fmt = item
            fmt = poison_fmt or clean_fmt
        else:
            label, value, fmt = item
        row = _apply_data_row(ws, row, label, value, fmt)
    row += 1

    # Income Statement
    row = _apply_section_header(ws, row, "Income Statement (FY2025)")
    section = data["income_statement"]
    for item in section:
        if is_poisoned:
            label, value, clean_fmt, poison_fmt = item
            fmt = poison_fmt or clean_fmt
        else:
            label, value, fmt = item
        is_sub = label in ("D&A",)
        row = _apply_data_row(ws, row, label, value, fmt, is_sublabel=is_sub)
    row += 1

    # Balance Sheet
    row = _apply_section_header(ws, row, "Balance Sheet (as of Dec 31, 2025)")
    section = data["balance_sheet"]
    for item in section:
        if is_poisoned:
            label, value, clean_fmt, poison_fmt = item
            fmt = poison_fmt or clean_fmt
        else:
            label, value, fmt = item
        row = _apply_data_row(ws, row, label, value, fmt)
    row += 1

    # Key Metrics
    row = _apply_section_header(ws, row, "Key Financial Metrics")
    section = data["key_metrics"]
    for item in section:
        if is_poisoned:
            label, value, clean_fmt, poison_fmt = item
            fmt = poison_fmt or clean_fmt
        else:
            label, value, fmt = item
        row = _apply_data_row(ws, row, label, value, fmt)

    # Footer
    row += 2
    ws.cell(row=row, column=1, value="Source: Company management; unaudited").font = Font(
        name="Calibri", size=9, italic=True, color="999999"
    )

    ws.sheet_properties.pageSetUpPr = None
    ws.protection.sheet = False

    return wb


def generate_clean():
    wb = _build_workbook(CLEAN_DATA, is_poisoned=False)
    path = os.path.join(EXAMPLES_DIR, "financials_clean.xlsx")
    wb.save(path)
    print(f"  Created {path}")
    return path


def generate_poisoned():
    wb = _build_workbook(POISONED_DATA, is_poisoned=True)
    path = os.path.join(EXAMPLES_DIR, "financials_poisoned.xlsx")
    wb.save(path)
    print(f"  Created {path}")
    return path


def main():
    os.makedirs(EXAMPLES_DIR, exist_ok=True)
    print("Generating XLSX variants...")
    generate_clean()
    generate_poisoned()
    print("\nDone. Files in examples/")

    print("\n--- POISONED SUMMARY ---")
    print("What Excel displays vs. what extractors will read:\n")
    print(f"  {'Metric':<30s} {'Display':>18s} {'Raw Value':>18s}")
    print(f"  {'-'*30} {'-'*18} {'-'*18}")
    for section in ("income_statement", "balance_sheet", "key_metrics"):
        for item in POISONED_DATA[section]:
            label, raw, clean_fmt, poison_fmt = item
            if not label or raw is None or not poison_fmt:
                continue
            display = poison_fmt.replace('"', '')
            print(f"  {label:<30s} {display:>18s} {str(raw):>18s}")


if __name__ == "__main__":
    main()
